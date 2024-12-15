import openpyxl
import os
import numpy as np
import sys
import count_pairs_tool
from collections import defaultdict

def _read_txt_strip_(url):
    file = open(url, 'r', encoding='utf-8')
    lines = file.readlines()
    file.close()
    return [line.strip() for line in lines]


def extract_results_from_file(result_path, pt, metric, language_num, language_dict, keys_map):
    data = _read_txt_strip_(os.path.join(result_path, f"{pt}.{metric}"))
    score_list = np.zeros((language_num, language_num))
    idx_i, idx_j = '', ''
    flag = False
    for row in data:
        if "-" in row and len(row) < 8:
            tmp = row.split("-")
            if tmp[0] in language_dict.keys() and tmp[1] in language_dict.keys():
                flag = True
            else:
                flag = False
            if flag == False: continue
            idx_i, idx_j = language_dict[tmp[0]], language_dict[tmp[1]]
        if f"{keys_map[metric]}" in row:
            if flag == False: continue
            if metric != "comet":
                score = float(row.split(":")[2].split(",")[0])
            else:
                score = float(row.split(":")[1])
            score = round(score, 2)
            score_list[idx_i][idx_j] = score
    return score_list

def make_bar(sheet, row_num, metric, scores_class, languages_families):
    start_position=1
    sheet.cell(row=row_num, column=(start_position+1)).value = metric
    # overview
    sheet.cell(row=(row_num + 1), column=start_position).value = "supervised"
    sheet.cell(row=(row_num + 1), column=(start_position+1)).value = "zero"
    sheet.cell(row=(row_num + 1), column=(start_position+2)).value = "averaged"
    sheet.cell(row=(row_num + 2), column=start_position).value = round(scores_class[f"supervised"], 2)
    sheet.cell(row=(row_num + 2), column=(start_position+1)).value = round(scores_class[f"zero"], 2)
    sheet.cell(row=(row_num + 2), column=(start_position+2)).value = round(scores_class[f"averaged"], 2)

    # detailed
    families = list(languages_families.keys())

    count = 0
    for family in families:
        sheet.cell(row=(row_num + 3), column=(start_position + count)).value = f"{family}2x"
        sheet.cell(row=(row_num + 4), column=(start_position + count)).value = round(scores_class[f"{family}2x"], 2)
        count = count + 1
        sheet.cell(row=(row_num + 3), column=(start_position + count)).value = f"x2{family}"
        sheet.cell(row=(row_num + 4), column=(start_position + count)).value = round(scores_class[f"x2{family}"], 2)
        count = count + 1

    # count by target
    count = 0
    for src_family in families:
        for tgt_family in families:
            if not (src_family == "English" and tgt_family == "English"):
                sheet.cell(row=(row_num + 5), column=(start_position + count)).value = f"{src_family}2{tgt_family}"
                sheet.cell(row=(row_num + 6), column=(start_position + count)).value = round(scores_class[f"{src_family}2{tgt_family}"], 2)
                count = count + 1




def count_results_by_class(scores, language_dict, bridge_languages, languages_families, metric):
    families = list(languages_families.keys())
    language_to_family = {lang: family for family, langs in languages_families.items() for lang in langs}
    special_pairs = count_pairs_tool.special_pairs
    
    # 初始化统计和计数字典
    stats = defaultdict(float)
    counts = defaultdict(int)
    
    # 准备统计类别的键
    for src_family in families:
        for tgt_family in families:
            if not (src_family == "English" and tgt_family == "English"):
                stats[f"{src_family}2{tgt_family}"] = 0.0
                counts[f"{src_family}2{tgt_family}"] = 0
        stats[f"x2{src_family}"] = 0.0
        counts[f"x2{src_family}"] = 0
        stats[f"{src_family}2x"] = 0.0
        counts[f"{src_family}2x"] = 0
    
    stats["supervised"] = 0.0
    counts["supervised"] = 0
    stats["zero"] = 0.0
    counts["zero"] = 0
    stats["averaged"] = 0.0
    counts["averaged"] = 0
    
    # 遍历语言对
    for src, src_id in language_dict.items():
        for tgt, tgt_id in language_dict.items():
            if src == tgt:
                continue
            tmp_score = scores[src_id, tgt_id]
            
            # 更新监督学习和零样本学习的统计和计数
            if src in bridge_languages or tgt in bridge_languages or f"{src}-{tgt}" in special_pairs or f"{tgt}-{src}" in special_pairs:
                stats["supervised"] += tmp_score
                counts["supervised"] += 1
            else:
                stats["zero"] += tmp_score
                counts["zero"] += 1
            
            # 更新基于语言族的统计和计数
            src_family = language_to_family.get(src)
            tgt_family = language_to_family.get(tgt)
            if src_family and tgt_family:
                key = f"{src_family}2{tgt_family}"
                stats[key] += tmp_score
                counts[key] += 1
                stats[f"x2{tgt_family}"] += tmp_score
                counts[f"x2{tgt_family}"] += 1
                stats[f"{src_family}2x"] += tmp_score
                counts[f"{src_family}2x"] += 1

            # 更新总的统计和计数
            stats["averaged"] += tmp_score
            counts["averaged"] += 1
    
    # 计算平均值
    average_stats = {key: (stats[key] / counts[key]) if counts[key] > 0 else 0 for key in stats.keys()}
    return average_stats


def make_detailed_table(sheet, start_row, metric, scores, language_sequence):
    sheet.cell(row=start_row, column=1).value = metric
    sheet.cell(row=(start_row + 1), column=1).value = "src"
    sheet.cell(row=start_row, column=2).value = "tgt"
    language_num = len(language_sequence)
    for i in range(language_num):
        sheet.cell(row=(start_row + 1), column=(2 + i)).value = language_sequence[i]
        sheet.cell(row=(start_row + 2 + i), column=1).value = language_sequence[i]
        for j in range(language_num):
            if scores[i][j] == 0: continue
            sheet.cell(row=(start_row + 2) + i, column=(2 + j)).value = scores[i][j]



def mk_table(work_path, experiment_name, experiment_id, pt, metric_list, language_sequence, language_dict, bridge_languages, languages_families, keys_map):
    result_path = os.path.join(work_path,"results", experiment_name, experiment_id)
    language_num = len(language_sequence)
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet(index=0, title="sheet1")
    for i in range(len(metric_list)):
        metric = metric_list[i]
        scores = extract_results_from_file(result_path, pt, metric, language_num, language_dict, keys_map)
        scores_class = count_results_by_class(scores, language_dict, bridge_languages, languages_families, metric)
        make_bar(sheet, (i * 7) + 1, metric, scores_class, languages_families)
        make_detailed_table(sheet, (i * (len(language_sequence) + 3) + (8 * len(metric_list))), metric, scores, language_sequence)
    save_dir = os.path.join("tables", experiment_name, experiment_id)
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
    wb.save(os.path.join(save_dir, f"{pt}.xlsx"))


language_sequence = count_pairs_tool.languages
bridge_languages = count_pairs_tool.all_bridge_languages
languages_families = count_pairs_tool.languages_families

language_dict = {key: index for index, key in enumerate(language_sequence)}

work_path = sys.argv[1]
experiment_name = sys.argv[2]
experiment_id = sys.argv[3]
pt = sys.argv[4]

keys_map = {
    "spbleu": "score",
    "chrf": "score",
    "comet": "Score",
}
metric_list = list(keys_map.keys())

mk_table(work_path, experiment_name, experiment_id, pt, metric_list, language_sequence, language_dict, bridge_languages, languages_families, keys_map)