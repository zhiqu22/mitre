import openpyxl
import os
import numpy as np
import sys
from collections import defaultdict


offtarget_languages = [
    "en", "bg", "da", "es", "uk", "hi", "ro", "de", "cs", "pt",
    "nl", "mr", "ur", "sv", "gu", "ar", "fr", "ru", "it", "pl",
    "he", "kn", "bn", "be", "mt", "am", "is", "sd"
]
comet_languages = [
    "en", "bg", "so", "ca", "da", "be", "bs", "es", "uk", "am", "hi", "ro",
    "no", "de", "cs", "pt", "nl", "mr", "is", "ne", "ur",
    "ha", "sv", "gu", "ar", "fr", "ru", "it", "pl", "sr", "sd", "he", "af", "kn", "bn",
]

def _read_txt_strip_(url):
    file = open(url, 'r', encoding='utf-8')
    lines = file.readlines()
    file.close()
    return [line.strip() for line in lines]


def extract_results_from_file(result_path, pt, metric, language_num, language_dict, keys_map):
    data = _read_txt_strip_(os.path.join(result_path, f"{pt}.{metric}"))
    score_list = np.zeros((language_num, language_num))
    idx_i, idx_j = '', ''
    for row in data:
        if "-" in row and len(row) < 8:
            tmp = row.split("-")
            idx_i, idx_j = language_dict[tmp[0]], language_dict[tmp[1]]
        if f"{keys_map[metric]}" in row:
            if metric != "comet" and  metric != "offtarget":
                score = float(row.split(":")[2].split(",")[0])
            else:
                score = float(row.split(":")[1])
            score = round(score, 2)
            score_list[idx_i][idx_j] = score
    return score_list

def make_bar(sheet, row_num, metric, scores, scores_class):
    start_position=1
    sheet.cell(row=row_num, column=(start_position+1)).value = metric
    # overview
    sheet.cell(row=(row_num + 1), column=start_position).value = "en2m"
    sheet.cell(row=(row_num + 1), column=(start_position+1)).value = "m2en"
    sheet.cell(row=(row_num + 1), column=(start_position+2)).value = "supervised"
    sheet.cell(row=(row_num + 1), column=(start_position+3)).value = "zero"
    sheet.cell(row=(row_num + 1), column=(start_position+4)).value = "averaged"
    en2m, m2en = np.sum(scores[0, :]) / np.count_nonzero(scores[0, :]), np.sum(scores[:, 0]) / np.count_nonzero(scores[:, 0])
    sheet.cell(row=(row_num + 2), column=start_position).value = round(en2m, 2)
    sheet.cell(row=(row_num + 2), column=(start_position+1)).value = round(m2en, 2)
    sheet.cell(row=(row_num + 2), column=(start_position+2)).value = round((m2en + en2m) / 2, 2)
    sheet.cell(row=(row_num + 2), column=(start_position+3)).value = round(np.sum(scores[1:, 1:]) / np.count_nonzero(scores[1:, 1:]), 2)
    sheet.cell(row=(row_num + 2), column=(start_position+4)).value = round(np.sum(scores) / np.count_nonzero(scores), 2)

    # detailed
    count = 0
    for src in ["h","m","l","e"]:
        for tgt in ["h","m","l","e"]:
            sheet.cell(row=(row_num + 3), column=(start_position + count)).value = f"{src}2{tgt}"
            sheet.cell(row=(row_num + 4), column=(start_position + count)).value = round(scores_class[f"{src}2{tgt}"], 2)
            count = count + 1
    # count by target
    count = 0
    for tgt in ["h","m","l","e"]:
        mean = sum([scores_class[f"{src}2{tgt}"] for src in ["h","m","l","e"]]) / 4 
        sheet.cell(row=(row_num + 5), column=(start_position + count)).value = f"~2{tgt}"
        sheet.cell(row=(row_num + 6), column=(start_position + count)).value = round(mean, 2)
        count = count + 1
    
    # count by source
    count = 0
    for src in ["h","m","l","e"]:
        mean = sum([scores_class[f"{src}2{tgt}"] for tgt in ["h","m","l","e"]]) / 4 
        sheet.cell(row=(row_num + 7), column=(start_position + count)).value = f"{src}2~"
        sheet.cell(row=(row_num + 8), column=(start_position + count)).value = round(mean, 2)
        count = count + 1

def count_results_by_class(scores, language_dict, language_classification, metric):
    stats_sum = defaultdict(float)
    stats_count = defaultdict(int)

    for src_class in language_classification:
        for tgt_class in language_classification:
            key = f"{src_class}2{tgt_class}"
            stats_sum[key] = 0.0
            stats_count[key] = 0

    for src_class, src_langs in language_classification.items():
        for tgt_class, tgt_langs in language_classification.items():
            key = f"{src_class}2{tgt_class}"
            for src_lang in src_langs:
                for tgt_lang in tgt_langs:
                    if src_lang == tgt_lang:
                        continue
                    if metric == "offtarget" and (src_lang not in offtarget_languages and tgt_lang not in offtarget_languages):
                        continue
                    if metric == "comet" and (src_lang not in comet_languages and tgt_lang not in comet_languages):
                        continue
                    src_id = language_dict[src_lang]
                    tgt_id = language_dict[tgt_lang]
                    score = scores[src_id, tgt_id]
                    
                    stats_sum[key] += score
                    stats_count[key] += 1

    average_stats = {
        key: (stats_sum[key] / stats_count[key]) if stats_count[key] > 0 else 0
        for key in stats_sum.keys()
    }
    return average_stats


def make_detailed_table(sheet, start_row, metric, scores, language_sequence):
    sheet.cell(row=start_row, column=1).value = metric
    sheet.cell(row=(start_row + 1), column=1).value = "src"
    sheet.cell(row=start_row, column=2).value = "tgt"
    language_num = len(language_sequence)
    for i in range(language_num):
        sheet.cell(row=(start_row + 1), column=(2 + i)).value = language_sequence[i]
        sheet.cell(row=(start_row + 2 + i), column=1).value = language_sequence[i]
        for j in range(language_num):
            if scores[i][j] == 0: continue
            sheet.cell(row=(start_row + 2) + i, column=(2 + j)).value = scores[i][j]



def mk_table(work_path, experiment_name, experiment_id, pt, metric_list, language_sequence, language_dict, language_classification, keys_map):
    result_path = os.path.join(work_path,"results", experiment_name, experiment_id)
    language_num = len(language_sequence)
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet(index=0, title="sheet1")
    for i in range(len(metric_list)):
        metric = metric_list[i]
        scores = extract_results_from_file(result_path, pt, metric, language_num, language_dict, keys_map)
        scores_class = count_results_by_class(scores, language_dict, language_classification, metric)
        make_bar(sheet, (i * 9) + 1, metric, scores, scores_class)
        make_detailed_table(sheet, (i * (len(language_sequence) + 3) + (len(metric_list) * 10)), metric, scores, language_sequence)
    save_dir = os.path.join("tables", experiment_name, experiment_id)
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
    wb.save(os.path.join(save_dir, f"{pt}.xlsx"))

language_sequence = [
    "en", "bg", "so", "ca", "da", "be", "bs", "mt", "es", "uk", "am", "hi", "ro",
    "no", "ti", "de", "cs", "lb", "pt", "nl", "mr", "is", "ne", "ur", "oc",
    "ast", "ha", "sv", "kab", "gu", "ar", "fr", "ru", "it", "pl", "sr", "sd", "he", "af", "kn", "bn",
]
language_classification = {
    "h": ["en", "de", "nl", "fr", "es", "ru", "cs", "hi", "bn", "ar", "he"],
    "m": ["sv", "da", "it", "pt", "pl", "bg", "kn", "mr", "mt", "ha"],
    "l": ["af", "lb", "ro", "oc", "uk", "sr", "sd", "gu", "ti", "am"],
    "e": ["no", "is", "ast", "ca", "be", "bs", "ne", "ur", "kab", "so"],
}
language_dict = {key: index for index, key in enumerate(language_sequence)}

metric_list = ["spbleu", "chrf", "comet", "offtarget"]
work_path = sys.argv[1]
experiment_name = sys.argv[2]
experiment_id = sys.argv[3]
pt = sys.argv[4]
keys_map = {
    "spbleu": "score",
    "chrf": "score",
    "comet": "Score",
    "offtarget": "Ratio",
}

mk_table(work_path, experiment_name, experiment_id, pt, metric_list, language_sequence, language_dict, language_classification, keys_map)